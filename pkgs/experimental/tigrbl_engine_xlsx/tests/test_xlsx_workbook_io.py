from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from tigrbl_engine_xlsx import xlsx_engine


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "widgets"
    sheet.append(["id", "name"])
    sheet.append(["1", "a"])
    workbook.save(path)


def test_xlsx_workbook_table_sheet_and_save_support(tmp_path: Path) -> None:
    path = tmp_path / "demo.xlsx"
    _make_workbook(path)

    _, session_factory = xlsx_engine(mapping={"path": str(path), "pk": "id"})
    session = session_factory()

    wb = session.workbook()
    assert load_workbook(str(path)).sheetnames == ["widgets"]
    assert wb["widgets"].title == "widgets"

    clone = tmp_path / "clone.xlsx"
    wb.save(clone)
    assert clone.exists()

    session._cat.tables["widgets"].append({"id": "2", "name": "b"})
    session._persist_workbook()

    reloaded = load_workbook(str(path))
    rows = list(reloaded["widgets"].iter_rows(values_only=True))
    assert rows[1:] == [("1", "a"), ("2", "b")]


def test_xlsx_commit_uses_atomic_replace(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "atomic.xlsx"
    _make_workbook(path)
    _, session_factory = xlsx_engine(mapping={"path": str(path), "pk": "id"})
    session = session_factory()

    calls: list[tuple[str, str]] = []
    real_replace = __import__("os").replace

    def capture_replace(src: str, dst: str) -> None:
        calls.append((src, dst))
        real_replace(src, dst)

    monkeypatch.setattr("tigrbl_engine_xlsx.session.os.replace", capture_replace)

    session._cat.tables["widgets"].append({"id": "3", "name": "c"})
    session._persist_workbook()

    assert path.exists()
    assert len(calls) == 1
    assert calls[0][1] == str(path)
    assert Path(calls[0][0]).name.startswith(".tmp_")

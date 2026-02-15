from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column

from tigrbl_engine_xlsx import xlsx_engine


class Base(DeclarativeBase):
    pass


class Widget(Base):
    __tablename__ = "widgets"

    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str]


@pytest.mark.asyncio
async def test_xlsx_session_commit_persists_rows_atomically(tmp_path: Path) -> None:
    path = tmp_path / "widgets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "widgets"
    sheet.append(["id", "name"])
    sheet.append([1, "a"])
    workbook.save(path)

    _, make_session = xlsx_engine(mapping={"path": str(path), "pk": "id"})
    session = make_session()

    await session.begin()
    session.add(Widget(id=2, name="b"))
    await session.commit()

    wb = load_workbook(path)
    rows = list(wb["widgets"].iter_rows(values_only=True))
    assert rows == [("id", "name"), (1, "a"), (2, "b")]

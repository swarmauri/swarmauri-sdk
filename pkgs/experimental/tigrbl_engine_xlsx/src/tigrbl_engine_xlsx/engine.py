from __future__ import annotations

from dataclasses import dataclass, field
import os
import tempfile
import threading
from typing import TYPE_CHECKING, Any, Callable

from openpyxl import Workbook, load_workbook

if TYPE_CHECKING:
    from .session import XlsxSession


@dataclass
class WorkbookCatalog:
    workbook: Workbook
    path: str
    pks: dict[str, str] = field(default_factory=dict)
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    table_ver: dict[str, int] = field(default_factory=dict)
    lock: threading.RLock = field(default_factory=threading.RLock)

    def get_live(self, name: str) -> list[dict[str, Any]]:
        self.tables.setdefault(name, [])
        self.table_ver.setdefault(name, 0)
        return self.tables[name]

    def bump(self, name: str) -> None:
        self.table_ver[name] = self.table_ver.get(name, 0) + 1


def _sheet_to_rows(workbook: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value) for value in values[0] if value is not None]
    rows: list[dict[str, Any]] = []
    for row_values in values[1:]:
        if all(value is None for value in row_values):
            continue
        row = {
            headers[index]: row_values[index]
            for index in range(min(len(headers), len(row_values)))
        }
        rows.append(row)
    return rows


def atomic_save_xlsx(workbook: Workbook, path: str) -> None:
    directory = os.path.dirname(path) or "."
    fd, tmp = tempfile.mkstemp(dir=directory, prefix=".tmp_", suffix=".xlsx")
    os.close(fd)
    try:
        workbook.save(tmp)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


@dataclass
class XlsxEngine:
    """Workbook-backed engine: workbook as DB, sheet names as table names."""

    path: str
    pk: str
    catalog: WorkbookCatalog


def xlsx_engine(
    *, mapping: dict[str, Any] | None = None, **_: Any
) -> tuple[XlsxEngine, Callable[[], XlsxSession]]:
    config = dict(mapping or {})
    path = config.get("path") or config.get("file")
    if not isinstance(path, str) or not path:
        raise TypeError("mapping['path'] (or 'file') must be a non-empty string")

    pk = config.get("pk", "id")
    if not isinstance(pk, str) or not pk:
        raise TypeError("mapping['pk'] must be a non-empty string")

    workbook = load_workbook(path)
    tables = {sheet: _sheet_to_rows(workbook, sheet) for sheet in workbook.sheetnames}
    if not tables:
        raise ValueError("workbook must contain at least one sheet")

    catalog = WorkbookCatalog(
        workbook=workbook,
        path=path,
        tables=tables,
        pks={sheet: pk for sheet in tables},
    )
    engine = XlsxEngine(path=path, pk=pk, catalog=catalog)

    from .session import XlsxSession

    def session_factory() -> XlsxSession:
        return XlsxSession(engine)

    return engine, session_factory


def xlsx_capabilities() -> dict[str, object]:
    return {
        "files": ["xlsx"],
        "workbook_database": True,
        "multi_table": True,
        "transactions": True,
        "dialect": "xlsx",
    }

from __future__ import annotations

import os
import tempfile
import uuid

from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Dict,
    List,
    Mapping,
    Optional,
    Sequence,
    Tuple,
)

from openpyxl.worksheet.worksheet import Worksheet

try:
    from tigrbl.session.base import TigrblSessionBase
except Exception:
    from abc import ABC, abstractmethod

    class TigrblSessionBase(ABC):  # pragma: no cover - fallback
        def __init__(self, spec=None):
            self._spec = spec

        @abstractmethod
        def _add_impl(self, obj): ...

        @abstractmethod
        async def _delete_impl(self, obj): ...

        @abstractmethod
        async def _flush_impl(self): ...

        @abstractmethod
        async def _refresh_impl(self, obj): ...

        @abstractmethod
        async def _get_impl(self, model, ident): ...

        @abstractmethod
        async def _execute_impl(self, stmt): ...

        @abstractmethod
        async def _tx_begin_impl(self): ...

        @abstractmethod
        async def _tx_commit_impl(self): ...

        @abstractmethod
        async def _tx_rollback_impl(self): ...

        @abstractmethod
        async def _close_impl(self): ...


try:
    from tigrbl.session.spec import SessionSpec
except Exception:

    class SessionSpec:  # pragma: no cover - fallback
        def __init__(self, isolation=None, read_only=None):
            self.isolation = isolation
            self.read_only = read_only


try:
    from tigrbl.core.crud.helpers.model import _single_pk_name, _model_columns
except Exception:

    def _single_pk_name(model):  # pragma: no cover - fallback
        return "id"

    def _model_columns(model):  # pragma: no cover - fallback
        return getattr(model, "__annotations__", {}) or {"id": int}


try:
    from tigrbl.core.crud.helpers import NoResultFound
except Exception:

    class NoResultFound(Exception):
        pass


if TYPE_CHECKING:
    from .engine import WorkbookCatalog


class _ScalarResult:
    def __init__(self, items: Sequence[Any]) -> None:
        self._items = list(items)

    def scalars(self) -> "_ScalarResult":
        return self

    def all(self) -> List[Any]:
        return list(self._items)

    def scalar_one(self) -> Any:
        if len(self._items) != 1:
            raise NoResultFound("expected exactly one row")
        return self._items[0]


class _ExecuteResult(_ScalarResult):
    rowcount: int = 0


class XlsxSession(TigrblSessionBase):
    """Native transaction session over workbook sheets as tables."""

    def __init__(
        self, catalog: "WorkbookCatalog", spec: Optional[SessionSpec] = None
    ) -> None:
        super().__init__(spec)
        self._cat = catalog
        self._snap: Dict[str, list[dict[str, Any]]] = {}
        self._snap_ver: Dict[str, int] = {}
        self._puts: Dict[Tuple[type, Any], Dict[str, Any]] = {}
        self._dels: set[Tuple[type, Any]] = set()
        self._tracked: Dict[Tuple[type, Any], Any] = {}

    def workbook(self):
        return self._cat.workbook

    def sheet(self, name: str) -> Worksheet:
        return self._cat.workbook[name]

    def table(self, name: str) -> list[dict[str, Any]]:
        rows = [dict(row) for row in self._cat.get_live(name)]
        pk = self._pk_of(name)
        by_pk = {row.get(pk): row for row in rows}

        for (model, ident), row in self._puts.items():
            if self._table(model) == name:
                by_pk[ident] = dict(row)

        for model, ident in self._dels:
            if self._table(model) == name:
                by_pk.pop(ident, None)

        return list(by_pk.values())

    async def run_sync(self, fn: Callable[[Any], Any]) -> Any:
        out = fn(self)
        return await out if hasattr(out, "__await__") else out

    async def _tx_begin_impl(self) -> None:
        self._snap.clear()
        self._snap_ver.clear()
        self._puts.clear()
        self._dels.clear()
        self._tracked.clear()

    async def _tx_commit_impl(self) -> None:
        iso = (self._spec.isolation if self._spec else None) or "read_committed"
        if iso in ("repeatable_read", "snapshot", "serializable"):
            for table, version in self._snap_ver.items():
                if self._cat.table_ver.get(table, 0) != version:
                    raise RuntimeError(f"transaction conflict on table '{table}'")

        with self._cat.lock:
            dels_by_table: Dict[str, List[Any]] = {}
            for model, ident in self._dels:
                dels_by_table.setdefault(self._table(model), []).append(ident)
            for table, idents in dels_by_table.items():
                pk = self._pk_of(table)
                live = self._cat.get_live(table)
                self._cat.tables[table] = [
                    row for row in live if row.get(pk) not in idents
                ]
                self._cat.bump(table)

            puts_by_table: Dict[str, List[Dict[str, Any]]] = {}
            for (model, _), row in self._puts.items():
                puts_by_table.setdefault(self._table(model), []).append(dict(row))
            for table, rows in puts_by_table.items():
                pk = self._pk_of(table)
                live = self._cat.get_live(table)
                live_by_pk = {row.get(pk): dict(row) for row in live}
                for row in rows:
                    if pk not in row:
                        raise RuntimeError(f"missing pk '{pk}' for table '{table}'")
                    live_by_pk[row[pk]] = row
                self._cat.tables[table] = list(live_by_pk.values())
                self._cat.bump(table)

            self._persist_workbook()

        self._puts.clear()
        self._dels.clear()
        self._tracked.clear()

    async def _tx_rollback_impl(self) -> None:
        self._snap.clear()
        self._snap_ver.clear()
        self._puts.clear()
        self._dels.clear()
        self._tracked.clear()

    @staticmethod
    def _pk_default(model: type, pk: str) -> Any:
        table = getattr(model, "__table__", None)
        if table is None:
            return None
        try:
            column = table.columns.get(pk)
        except Exception:
            return None
        if column is None:
            return None
        default = getattr(column, "default", None)
        if default is None:
            return None
        arg = getattr(default, "arg", None)
        if callable(arg):
            try:
                return arg()
            except TypeError:
                return arg(None)
        return arg

    def _add_impl(self, obj: Any) -> Any:
        model = obj.__class__
        pk = _single_pk_name(model)
        ident = getattr(obj, pk)
        if ident is None:
            ident = self._pk_default(model, pk)
            if ident is not None:
                setattr(obj, pk, ident)
        if ident is None:
            raise ValueError(f"primary key {pk!r} must be set")
        row = {column: getattr(obj, column, None) for column in _model_columns(model)}
        self._puts[(model, ident)] = row
        self._dels.discard((model, ident))
        self._tracked[(model, ident)] = obj
        return None

    async def _delete_impl(self, obj: Any) -> None:
        model = obj.__class__
        pk = _single_pk_name(model)
        ident = getattr(obj, pk)
        self._puts.pop((model, ident), None)
        self._dels.add((model, ident))
        self._tracked.pop((model, ident), None)

    async def _flush_impl(self) -> None:
        for (model, ident), obj in list(self._tracked.items()):
            if (model, ident) in self._dels:
                continue
            row = {
                column: getattr(obj, column, None) for column in _model_columns(model)
            }
            self._puts[(model, ident)] = row
        return

    async def _refresh_impl(self, obj: Any) -> None:
        pk = _single_pk_name(obj.__class__)
        ident = getattr(obj, pk)
        fresh = await self._get_impl(obj.__class__, ident)
        if fresh is None:
            return
        for column in _model_columns(obj.__class__):
            setattr(obj, column, getattr(fresh, column, None))

    async def _get_impl(self, model: type, ident: Any) -> Any | None:
        tracked = self._tracked.get((model, ident))
        if tracked is not None:
            return tracked
        row = self._puts.get((model, ident))
        if row is not None:
            return self._inflate_tracked(model, ident, row)
        if (model, ident) in self._dels:
            return None
        table_rows = self._rows_for(model)
        pk = _single_pk_name(model)
        for table_row in table_rows:
            if table_row.get(pk) == ident:
                return self._inflate_tracked(model, ident, table_row)
        return None

    async def _execute_impl(self, stmt: Any) -> Any:
        kind = type(stmt).__name__.lower()
        if "select" in kind:
            model, where, order, limit, offset = self._decompose_select(stmt)
            items = self._scan_model(model)
            items = [obj for obj in items if self._matches_obj(obj, where)]
            return _ExecuteResult(self._order_slice(items, order, limit, offset))
        if "delete" in kind:
            model, where = self._decompose_delete(stmt)
            items = self._scan_model(model)
            items = [obj for obj in items if self._matches_obj(obj, where)]
            for obj in items:
                pk = _single_pk_name(model)
                ident = getattr(obj, pk)
                self._puts.pop((model, ident), None)
                self._dels.add((model, ident))
                self._tracked.pop((model, ident), None)
            result = _ExecuteResult([])
            result.rowcount = len(items)
            return result
        raise NotImplementedError(f"Unsupported statement: {type(stmt)}")

    async def _close_impl(self) -> None:
        return

    def _table(self, model: type) -> str:
        return getattr(model, "__tablename__", None) or model.__name__

    def _pk_of(self, table: str) -> str:
        if table in self._cat.pks:
            return self._cat.pks[table]
        raise RuntimeError(f"primary key for table '{table}' is unknown")

    def _rows_for(self, model: type) -> list[dict[str, Any]]:
        table = self._table(model)
        iso = (self._spec.isolation if self._spec else None) or "read_committed"
        if (
            iso in ("repeatable_read", "snapshot", "serializable")
            and table not in self._snap
        ):
            live = self._cat.get_live(table)
            self._snap[table] = [dict(row) for row in live]
            self._snap_ver[table] = self._cat.table_ver.get(table, 0)
        return self._snap.get(table, self._cat.get_live(table))

    def _inflate(self, model: type, data: Mapping[str, Any]) -> Any:
        obj = model()
        for column in _model_columns(model):
            if column in data:
                setattr(obj, column, data[column])
        return obj

    def _inflate_tracked(self, model: type, ident: Any, data: Mapping[str, Any]) -> Any:
        obj = self._tracked.get((model, ident))
        if obj is None:
            obj = self._inflate(model, data)
            self._tracked[(model, ident)] = obj
        return obj

    def _scan_model(self, model: type) -> List[Any]:
        out = [self._inflate(model, row) for row in self._rows_for(model)]
        pk = _single_pk_name(model)
        by_id = {getattr(obj, pk): obj for obj in out}
        for (known_model, ident), row in self._puts.items():
            if known_model is model:
                by_id[ident] = self._inflate_tracked(model, ident, row)
        for ident, obj in list(by_id.items()):
            self._tracked[(model, ident)] = obj
        for known_model, ident in self._dels:
            if known_model is model:
                by_id.pop(ident, None)
        return list(by_id.values())

    def _persist_workbook(self) -> None:
        workbook = self._cat.workbook
        for table, rows in self._cat.tables.items():
            if table in workbook.sheetnames:
                sheet = workbook[table]
                sheet.delete_rows(1, sheet.max_row)
            else:
                sheet = workbook.create_sheet(table)
            columns: list[str] = []
            for row in rows:
                for key in row:
                    if key not in columns:
                        columns.append(key)
            if not columns:
                columns = [self._pk_of(table)]
            sheet.append(columns)
            for row in rows:
                sheet.append([self._excel_value(row.get(col)) for col in columns])
        self._atomic_save_workbook(workbook, self._cat.path)

    @staticmethod
    def _excel_value(value: Any) -> Any:
        if isinstance(value, uuid.UUID):
            return str(value)
        return value

    def _atomic_save_workbook(self, workbook: Any, path: str) -> None:
        directory = os.path.dirname(path) or "."
        fd, tmp = tempfile.mkstemp(dir=directory, prefix=".tmp_", suffix=".xlsx")
        os.close(fd)
        try:
            workbook.save(tmp)
            os.replace(tmp, path)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)

    def _decompose_select(
        self, stmt: Any
    ) -> Tuple[
        type,
        list[Tuple[str, str, Any]],
        list[Tuple[str, str]],
        Optional[int],
        Optional[int],
    ]:
        return (
            self._extract_model(stmt),
            self._extract_predicates(stmt),
            self._extract_order_by(stmt),
            self._extract_int(stmt, ["_limit", "_limit_clause", "limit"]),
            self._extract_int(stmt, ["_offset", "_offset_clause", "offset"]),
        )

    def _decompose_delete(self, stmt: Any) -> Tuple[type, list[Tuple[str, str, Any]]]:
        return self._extract_model(stmt), self._extract_predicates(stmt)

    def _extract_model(self, stmt: Any) -> type:
        descs = getattr(stmt, "column_descriptions", None) or []
        for desc in descs:
            entity = desc.get("entity") if isinstance(desc, dict) else None
            if isinstance(entity, type):
                return entity

        def _all_subclasses(base: type) -> list[type]:
            def _safe_subclasses(cls: type) -> list[type]:
                try:
                    return list(cls.__subclasses__())
                except TypeError:
                    return []

            out: list[type] = []
            stack = _safe_subclasses(base)
            while stack:
                cls = stack.pop()
                out.append(cls)
                stack.extend(_safe_subclasses(cls))
            return out

        def _find_by_table(name: str) -> type | None:
            for cls in _all_subclasses(object):
                if getattr(cls, "__tablename__", None) == name:
                    return cls
            return None

        for attr_name in ("_from_objects", "_froms", "froms"):
            value = getattr(stmt, attr_name, None)
            if value is not None:
                if isinstance(value, (list, tuple)) and not value:
                    continue
                table = value[0] if isinstance(value, (list, tuple)) else value
                name = getattr(table, "name", None)
                if isinstance(name, str):
                    found = _find_by_table(name)
                    if found is not None:
                        return found
        table = getattr(stmt, "table", None)
        name = getattr(table, "name", None)
        if isinstance(name, str):
            found = _find_by_table(name)
            if found is not None:
                return found

        rc = getattr(stmt, "_raw_columns", None) or getattr(stmt, "columns", None)
        if rc is not None:
            if isinstance(rc, (list, tuple)) and not rc:
                raise RuntimeError("Cannot resolve model from statement")
            entity = rc[0]
            table = getattr(entity, "table", None)
            name = getattr(table, "name", None)
            if isinstance(name, str):
                found = _find_by_table(name)
                if found is not None:
                    return found
        raise RuntimeError("Cannot resolve model from statement")

    def _extract_predicates(self, stmt: Any) -> list[Tuple[str, str, Any]]:
        where = getattr(stmt, "whereclause", None) or getattr(
            stmt, "_whereclause", None
        )
        if where is None:
            return []
        parts = (
            getattr(where, "clauses", None)
            or getattr(where, "get_children", lambda: [])()
        )
        nodes = list(parts) if parts else [where]
        out: list[Tuple[str, str, Any]] = []
        for node in nodes:
            left = getattr(node, "left", None)
            right = getattr(node, "right", None)
            op = getattr(node, "operator", None)
            if left is None or right is None:
                continue
            name = getattr(left, "key", None) or getattr(left, "name", None)
            if name is None:
                continue
            opname = getattr(op, "__name__", str(op))
            if "eq" in opname:
                value = (
                    getattr(right, "value", None)
                    if hasattr(right, "value")
                    else getattr(right, "literal", None)
                )
                out.append((str(name), "eq", value))
                continue
            right_clauses = getattr(right, "clauses", None)
            if right_clauses is not None and "in" in opname:
                vals = [
                    getattr(lit, "value", None)
                    if hasattr(lit, "value")
                    else getattr(lit, "literal", None)
                    for lit in right_clauses
                ]
                out.append((str(name), "in", vals))
        return out

    def _extract_order_by(self, stmt: Any) -> list[Tuple[str, str]]:
        order = getattr(stmt, "_order_by_clause", None)
        if order is None:
            order = getattr(stmt, "_order_by_clauses", None)
        if order is None:
            return []
        clauses = getattr(order, "clauses", None) or order
        clauses = clauses if isinstance(clauses, (list, tuple)) else [clauses]
        for ob in clauses:
            col = (
                getattr(ob, "element", None)
                or getattr(ob, "this", None)
                or getattr(ob, "expr", None)
            )
            name = getattr(col, "key", None) or getattr(col, "name", None)
            direction = "desc" if "desc" in type(ob).__name__.lower() else "asc"
            if name:
                return [(str(name), direction)]
        return []

    def _extract_int(self, stmt: Any, names: Sequence[str]) -> Optional[int]:
        for name in names:
            value = getattr(stmt, name, None)
            if value is None:
                continue
            try:
                return int(value)
            except Exception:
                lit = getattr(value, "value", None)
                if lit is not None:
                    try:
                        return int(lit)
                    except Exception:
                        pass
        return None

    def _matches_obj(self, obj: Any, where: list[Tuple[str, str, Any]]) -> bool:
        for name, op, value in where:
            datum = getattr(obj, name, None)
            if op == "eq" and datum != value:
                return False
            if op == "in" and datum not in set(value):
                return False
        return True

    def _order_slice(
        self,
        items: List[Any],
        order: list[Tuple[str, str]],
        limit: Optional[int],
        offset: Optional[int],
    ) -> List[Any]:
        if order:
            col, direction = order[0]
            items.sort(
                key=lambda obj: getattr(obj, col, None), reverse=(direction == "desc")
            )
        if isinstance(offset, int):
            items = items[max(0, offset) :]
        if isinstance(limit, int):
            items = items[: max(0, limit)]
        return items
